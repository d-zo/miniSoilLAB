# -*- coding: utf-8 -*-
"""
xlshilfen.py   v0.3 (2019-11)
"""

# Copyright 2020-2021 Dominik Zobel.
# All rights reserved.
#
# This file is part of the miniSoilLAB package.
# miniSoilLAB is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# miniSoilLAB is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with miniSoilLAB. If not, see <http://www.gnu.org/licenses/>.


_mustervorlagen = None;


# -------------------------------------------------------------------------------------------------
def LeseXLSDaten(dateiname, ignoriere=['rohdaten']):
   """Lese eine Excel-Tabelle namens dateiname ein. Abhaengig von den verfuegbaren Vorlagen wird
   unterschieden, welche Art von Versuchsdaten in der Datei gespeichert sind und versucht, ebenjene
   einzulesen. Gibt bei Erfolg eine Datenstruktur mit den eingelesenen Daten unter dem Schluessel
   Versuchsart zurueck (bspw. als LoDi fuer eine Tabelle zur Bestimmung der lockersten/dichtesten
   Lagerung oder Triax-D-dicht fuer einen drainierten Triaxialversuch mit dicht im Dateinamen).
   Wenn die Datei ignoriert wird oder nicht gefunden/geparst werden kann, wird eine leere Struktur
   zurueckgegeben (eine Datei wird ignoriert, wenn der Name (mindestens) ein Element aus der Liste
   ignoriere enthaelt).
   """
   from os import path as os_path
   from .datenstruktur import Datenstruktur
   from .vorlagen import VorlagenMuster
   from .konstanten import basispfad
   from .kennwerte import Kennwertberechnungen
   #
   bodendaten = Datenstruktur();
   global _mustervorlagen;
   if (_mustervorlagen is None):
      _mustervorlagen = VorlagenMuster();
      _mustervorlagen.AusOrdnerEinlesen(ordner=basispfad);
   #
   datei = os_path.basename(dateiname).lower();
   if (any([ignoriermuster in datei for ignoriermuster in ignoriere])):
      print('# - LeseXLS: Ignoriere ' + dateiname);
   #
   # FIXME: "with open" oder aehnliche Konstruktion, die automatisch das workbook wieder schliesst
   workbook, tabellentyp = _OeffneXLSDatei(dateiname=dateiname);
   if (workbook is None):
      return bodendaten;
   #
   zielvorlage, refvorlage = _mustervorlagen.Pruefen(workbook=workbook, tabellentyp=tabellentyp);
   # Wenn irgendeine Vorlage erfuellt ist, dann lese die Werte fuer diese Vorlage ein
   if (zielvorlage is None):
      print('# - LeseXLS: Keine Vorlage passt fuer ' + dateiname);
      return bodendaten;
   #
   print('# - LeseXLS (' + zielvorlage + '): ' + dateiname);
   musterdaten = Datenstruktur();
   tabellenseiten = _mustervorlagen.Datenfelder(schluessel=zielvorlage);
   for tabellenname in tabellenseiten.keys():
      ParseXLSDaten(daten=musterdaten, workbook=workbook, tabellentyp=tabellentyp,
         tabellenname=tabellenname, vorlage=tabellenseiten[tabellenname]);
   #
   if (Kennwertberechnungen(daten=musterdaten, vorlage=refvorlage)):
      musterdaten.update([('Status', 'Einlesen erfolgreich')]);
   else:
      musterdaten.update([('Status', 'Einlesen nicht erfolgreich (Daten unvollstaendig/ungueltig)')]);
   #
   musterdaten.update([('Dateiname', dateiname)]);
   refname = refvorlage;
   if ((refvorlage == 'Triax-D') or (refvorlage == 'Oedo')):
      lagerung = '';
      if ('dicht' in datei):
         lagerung = '-dicht';
      elif ('locker' in datei):
         lagerung = '-locker';
      elif ((('-di' in datei) or ('_di' in datei)) and ('-lo' not in datei) and ('_lo' not in datei)):
         lagerung = '-dicht';
      elif ((('-lo' in datei) or ('_lo' in datei)) and ('-di' not in datei) and ('_di' not in datei)):
         lagerung = '-locker';
      else:
         try:
            bodenart = musterdaten['Bodenart'];
         except:
            bodenart = '';
         #
         if ('locker' in bodenart.lower()):
            lagerung = '-locker';
         elif ('dicht' in bodenart.lower()):
            lagerung = '-dicht';
         #
         if (lagerung == ''):
            print('# Warnung: Konnte Lagerung (locker/dicht) weder aus Dateinamen noch Bodenart bestimmen. Nehme dicht an.');
            lagerung = '-dicht';
      #
      refname = refvorlage + lagerung;
   #
   bodendaten.update([(refname, musterdaten)]);
   return bodendaten;
#


# -------------------------------------------------------------------------------------------------
def _OeffneXLSDatei(dateiname):
   """Oeffnet eine Excel-Datei und gibt das darin enthaltene workbook und den tabellentyp (xls/xlsx)
   zurueck.
   """
   tabellentyp = '';
   if (dateiname[-3:].lower() == 'xls'):
      tabellentyp = 'xls';
      try:
         import xlrd
      except:
         print('# Fehler: xlrd (und Abhaengigkeiten) zum Laden von xls-Dateien nicht gefunden');
         return [None, None];
      #
      try:
         workbook = xlrd.open_workbook(filename=dateiname);
      except:
         print('# Fehler: Konnte .xls-Datei ' + dateiname + ' nicht finden/laden');
         return [None, None];
   elif (dateiname[-4:].lower() == 'xlsx'):
      tabellentyp = 'xlsx';
      try:
         import openpyxl
      except:
         print('# Fehler: openpyxl (und Abhaengigkeiten) zum Laden von xlsx-Dateien nicht gefunden');
         return [None, None];
      #
      try:
         workbook = openpyxl.load_workbook(filename=dateiname, read_only=True);
      except:
         print('# Fehler: Konnte .xlsx-Datei ' + dateiname + ' nicht finden/laden');
         return [None, None];
   else:
      print('# Fehler: Dateityp nicht unterstuetzt');
      return [None, None];
   #
   return [workbook, tabellentyp];
#   


# -------------------------------------------------------------------------------------------------
def _LeseEintrag(tabellenseite, zeile, spalte, tabellentyp):
   """Lese aus der uebergebenen tabellenseite einer eingelesenen Excel-Tabelle Daten aus einer Zelle
   aus. Fuer diese Zelle muss ueber zeile und spalte die exakte Position angegeben werden.
   Zusaetzlich ist der ueber den tabellentyp (xls/xlsx) anzugeben, ob die Datei und deren Daten
   nach den alten Dateistandards oder den XML-basierten Dateistandards zu interpretieren sind.
   Gibt den Inhalt der Zelle als String zurueck. Falls die Zelle nicht gefunden werden konnte,
   wird None zurueckgegeben.
   """
   if (tabellentyp == 'xls'):
      if (zeile >= tabellenseite.nrows) or (spalte >= tabellenseite.ncols):
         return None;
      #
      try:
         zelle = tabellenseite.row(zeile)[spalte].value;
         # Excel-interne Fehlercodes: nichts zurueckgeben
         if (tabellenseite.row(zeile)[spalte].ctype == 5):
            return None;
         #
      except:
         print('# Fehler: Konnte Wert aus angegebener Zelle nicht auslesen');
         return -1;
   else:
      if (zeile > tabellenseite.max_row) or (spalte > tabellenseite.max_column):
         return None;
      #
      try:
         zelle = tabellenseite.cell(row=zeile+1, column=spalte+1).value;
      except:
         print('# Fehler: Konnte Wert aus angegebener Zelle nicht auslesen');
         return -1;
   #
   if (zelle is not None):
      zelle = str(zelle);
   #
   return zelle;
#


# -------------------------------------------------------------------------------------------------
def TabellenDimension(workbook, tabellenname, tabellentyp):
   """Lese aus dem uebergebenen workbook einer eingelesenen Excel-Tabelle aus, wieviele Zeilen
   und Spalten in der Tabelle namens tabellenname maximal verwendet sind. Zusaetzlich ist der ueber
   den tabellentyp (xls/xlsx) anzugeben, ob die Datei und deren Daten nach den alten Dateistandards
   oder den XML-basierten Dateistandards zu interpretieren sind. Gibt [zeilen, spalten] zurueck
   oder [None, None], falls die Tabelle nicht gefunden werden konnte.
   """
   if (tabellentyp == 'xls'):
      try:
         xlssheet = workbook.sheet_by_name(tabellenname);
      except:
         print('# Fehler: Konnte Tabellenseite >' + tabellenname + '< nicht laden');
         return [None, None];
      #
      return [xlssheet.nrows, xlssheet.ncols];
   else:
      try:
         xlsxsheet = workbook[tabellenname];
      except:
         print('# Fehler: Konnte Tabellenseite >' + tabellenname + '< nicht laden');
         return [None, None];
      #
      return [xlsxsheet.max_row, xlsxsheet.max_column];
#
   

# -------------------------------------------------------------------------------------------------
def VorhandeneTabellenseiten(workbook, tabellentyp):
   """Lese aus dem uebergebenen workbook einer eingelesenen Excel-Tabelle die Liste an vorhandenen
   Tabellenseiten aus. Gibt diese Liste zurueck.
   """
   if (tabellentyp == 'xls'):
      return workbook.sheet_names();
   else:
      return workbook.sheetnames;
#


# -------------------------------------------------------------------------------------------------
def TabellenseiteAusgeben(workbook, tabellenname, tabellentyp):
   """Gibt aus dem uebergebenen workbook einer eingelesenen Excel-Tabelle die Tabellenseite mit dem
   Namen tabellenname zurueck. Dazu ist der ueber den tabellentyp (xls/xlsx) anzugeben, ob die Datei
   und deren Daten nach den alten Dateistandards oder den XML-basierten Dateistandards zu
   interpretieren sind. Falls die Tabellenseite nicht existiert, wird None zurueckgegeben.
   """
   if (tabellentyp == 'xls'):
      try:
         xlssheet = workbook.sheet_by_name(tabellenname);
      except:
         print('# Fehler: Konnte Tabellenseite >' + tabellenname + '< nicht laden');
         return None;
      #
      return xlssheet;
   else:
      try:
         xlsxsheet = workbook[tabellenname];
      except:
         print('# Fehler: Konnte Tabellenseite >' + tabellenname + '< nicht laden');
         return None;
      #
      return xlsxsheet;
#


# -------------------------------------------------------------------------------------------------
def EinzelEintragEinlesen(workbook, tabellenname, zeile, spalte, name, tabellentyp, inhalt, warnen=True):
   """Lese aus dem uebergebenen workbook einer eingelesenen Excel-Tabelle Daten aus einer Zelle aus.
   Fuer diese Zelle muss ueber tabellenname, zeile und spalte die exakte Position angegeben werden.
   Zusaetzlich ist der ueber den tabellentyp (xls/xlsx) anzugeben, ob die Datei und deren Daten
   nach den alten Dateistandards oder den XML-basierten Dateistandards zu interpretieren sind.
   Falls inhalt='zahl' wird der Zelleninhalt zu einem float umgewandelt und zurueckgegeben,
   andernfalls als String. Falls die Zelle nicht gefunden werden konnte, wird None zurueckgegeben.
   """
   tabellenseite = TabellenseiteAusgeben(workbook=workbook, tabellenname=tabellenname,
      tabellentyp=tabellentyp);
   if (tabellenseite is None):
      return None;
   else:
      return _EinzelEintragEinlesen(tabellenseite=tabellenseite, zeile=zeile, spalte=spalte,
         name=name, tabellentyp=tabellentyp, inhalt=inhalt, warnen=warnen);
#


# -------------------------------------------------------------------------------------------------
def _EinzelEintragEinlesen(tabellenseite, zeile, spalte, name, tabellentyp, inhalt, warnen=True):
   """Lese aus der uebergebenen tabellenseite einer eingelesenen Excel-Tabelle Daten aus einer Zelle
   aus. Fuer diese Zelle muss ueber zeile und spalte die exakte Position angegeben werden.
   Zusaetzlich ist der ueber den tabellentyp (xls/xlsx) anzugeben, ob die Datei und deren Daten
   nach den alten Dateistandards oder den XML-basierten Dateistandards zu interpretieren sind.
   Falls inhalt='zahl' wird der Zelleninhalt zu einem float umgewandelt und zurueckgegeben,
   andernfalls als String. Falls die Zelle nicht gefunden werden konnte, wird None zurueckgegeben.
   """
   neuerEintrag = _LeseEintrag(tabellenseite=tabellenseite, zeile=zeile, spalte=spalte,
      tabellentyp=tabellentyp);
   if (neuerEintrag == -1):
      return None;
   #
   if ((neuerEintrag is None) or (neuerEintrag == '')):
      if (warnen):
         print('# Warnung: ' + name + ' ungueltig');
      #
      return None;
   #
   if (inhalt == 'text'):
      rueckgabe = neuerEintrag;
   elif (inhalt == 'zahl'):
      try:
         rueckgabe = float(str(neuerEintrag).replace(',', '.'));
      except:
         print('# Warnung: Wert von ' + name + ' kann nicht in Zahl umgewandelt werden');
         rueckgabe = None;
   else:
      print('# Warnung: Zieltyp ' + inhalt + ' nicht unterstuetzt - kann nicht einlesen');
      rueckgabe = None;
   #
   return rueckgabe;
#


# -------------------------------------------------------------------------------------------------
def EinzelEintragHinzufuegen(datenbank, workbook, tabellenname, zeile, spalte, name, tabellentyp,
   inhalt, grenzen=None):
   """Lese aus dem uebergebenen workbook einer eingelesenen Excel-Tabelle Daten aus einer Zelle aus
   und speichere den Wert in datenbank. Fuer inhalt = 'zahl' kann auch eine Ueberpruefung der
   grenzen erfolgen (sofern diese nicht None sind). Ruft intern EinzelEintragEinlesen() auf und
   speichert den (erfolgreich) eingelesenen Zelleninhalt darin mit dem Schluessel name.
   """
   from .gleichungsloeser import WertInZulaessigemBereich
   #
   eingelesen = EinzelEintragEinlesen(workbook=workbook, tabellenname=tabellenname, zeile=zeile,
      spalte=spalte, name=name, tabellentyp=tabellentyp, inhalt=inhalt);
   if (eingelesen is not None):
      if (grenzen is None):
         datenbank.update([(name, eingelesen)]);
      else:
         if ((inhalt == 'zahl') and (grenzen is not None)):
            if (WertInZulaessigemBereich(name=name, liste=[eingelesen], minmax=grenzen)):
               datenbank.update([(name, eingelesen)]);
#


# -------------------------------------------------------------------------------------------------
def GelesenenEintragHinzufuegen(datenbank, name, daten, grenzen=None):
   """Speichere bereits eingelesene Inhalte aus daten in datenbank (unter dem Schluessel name).
   Fuer Zahlen kann auch eine Ueberpruefung der grenzen erfolgen (sofern diese nicht None sind).
   Nur erfolgreich gepruefte daten werden gespeichert.
   """
   from .gleichungsloeser import WertInZulaessigemBereich
   #
   if (grenzen is None):
      datenbank.update([(name, daten)]);
   else:
      if (isinstance(daten, list) or isinstance(daten, tuple)):
         grenze_eingehalten = WertInZulaessigemBereich(name=name, liste=daten, minmax=grenzen);
      else:
         grenze_eingehalten = WertInZulaessigemBereich(name=name, liste=[daten], minmax=grenzen);
      #
      if (grenze_eingehalten):
         datenbank.update([(name, daten)]);
#
 

# -------------------------------------------------------------------------------------------------
def ZellenbereichKoordinaten(workbook, tabellentyp, tabellenname, bereich):
   """Ermittelt, ob der uebergebene bereich (bspw. "D14 oder "G5:R5") fuer die Tabellenseite
   tabellenname im workbook des tabellentyps zulaessig ist, oder nicht. Wenn der Bereich zulaessig
   ist, werden die umgewandelten Koordinaten fuer jede Zelle [(zeile, spalte), (zeile, spalte), ...]
   zurueckgegeben, sonst [(None, None)].
   """
   maxzeile, maxspalte = TabellenDimension(workbook, tabellenname, tabellentyp);
   zellkoordinaten = [];
   gruppen = bereich.split(';');
   for einzelgruppe in gruppen:
      einzelbereiche = einzelgruppe.split(':');
      startzeile, startspalte = ZeileUndSpalteAusZellenbezeichnung(zellenname=einzelbereiche[0]);
      if ((startzeile < 0) or (startzeile > maxzeile) or (startspalte < 0) or (startspalte > maxspalte)):
         print('# Warnung: Zelle (' + einzelbereiche[0] + ') liegt nicht im gueltigen Tabellenbereich');
         return [(None, None)];
      #
      if (len(einzelbereiche) == 1):
         zellkoordinaten += [(startzeile, startspalte)];
         continue;
      elif (len(einzelbereiche) != 2):
         print('# Warnung: Ungueltige Intervalldefinition ' + einzelgruppe);
         return [(None, None)];
      #
      endzeile, endspalte = ZeileUndSpalteAusZellenbezeichnung(zellenname=einzelbereiche[1]);
      #
      if (endzeile == -1):
         endzeile = maxzeile;
      #
      if (endspalte == -1):
         endspalte = maxspalte;
      #
      if ((endzeile < 0) or (endzeile > maxzeile) or (endspalte < 0) or (endspalte > maxspalte)):
         print('# Warnung: Zelle (' + einzelbereiche[1] + ') liegt nicht im gueltigen Tabellenbereich');
         return [(None, None)];
      #
      if ((endzeile < startzeile) or (endspalte < startspalte)):
         print('# Warnung: Startzelle ist nicht links oberhalb Zielzelle in Bereich (' \
            + ':'.join(einzelbereiche) + ')');
         return [(None, None)];
      #
      if ((endzeile > startzeile) and (endspalte > startspalte)):
         print('# Warnung: Nur Zeilen- oder Spaltenbereich erlaubt (nicht beides)');
         return [(None, None)];
      #
      for idx_zeile in range(startzeile, endzeile+1):
         for idx_spalte in range(startspalte, endspalte+1):
            zellkoordinaten += [(idx_zeile, idx_spalte)];
   #
   return zellkoordinaten;
#


# -------------------------------------------------------------------------------------------------
def ZeileUndSpalteAusZellenbezeichnung(zellenname):
   """Ermitttle die Zeile und Spalte einer Zelle aus dem uebergebenem zellenname. Der Zellenname
   besteht aus einem oder mehreren Buchstaben fuer die Spalte und einer Zahl (ggfs. mit mehreren
   Stellen) fuer die Zeile. Gibt [zeile, spalte] zurueck.
   """
   idx_alpha = -1;
   idx_zahl = -1;
   letztes_alpha = False;
   letzte_zahl = False;
   ungueltig = False;
   # Eine korrekte Zellenbezeichnung faengt mit Buchstaben an und hoert mit Zahlen auf
   for idx, zeichen in enumerate(zellenname):
      if (zeichen.isalpha()):
         idx_alpha = idx;
         if (idx_zahl != -1):
            ungueltig = True;
            break;
      #
      elif (zeichen.isdigit()):
         idx_zahl = idx;
         if (idx_alpha == -1):
            ungueltig = True;
            break;
      elif (zeichen == '-'):
         if ((idx == 0) and (not letztes_alpha)):
            letztes_alpha = True;
         elif ((idx_alpha > -1) and (not letzte_zahl)):
            letzte_zahl = True;
         else:
            ungueltig = True;
            break;
      else:
         ungueltig = True;
         break;
   #
   if ((idx_alpha == -1) or (idx_zahl == -1)):
      ungueltig = True;
   #
   if (ungueltig):
      print('# Warnung: Ungueltige Zellenbezeichnung ' + zellenname);
      return [None, None];
   #
   # Kurzformen fuer Zeilenende und Spaltenende erlauben
   if (letzte_zahl):
      zeile = 0;
   else:
      zeile = int(zellenname[idx_alpha+1:]);
   #
   if (letztes_alpha):
      spalte = 0;
   else:
      spaltenname = zellenname[:idx_alpha+1].lower();
      refbuchstaben = 'abcdefghijklmnopqrstuvwxyz';
      spalte = 0;
      for buchstabe in spaltenname:
         spalte = 26*spalte + refbuchstaben.index(buchstabe) + 1;
   #
   return [zeile - 1, spalte - 1];
#


# -------------------------------------------------------------------------------------------------
def ParseXLSDaten(daten, workbook, tabellentyp, tabellenname, vorlage):
   """Ein (ggfs. verschachteltes) dict wird als vorlage uebergeben und abgearbeitet. Fuer jeden
   der Eintraege wird eine Struktur in daten angelegt und mit Werten aus der Tabelle tabellenname
   eines workbook des Typs tabellentyp aufgefuellt.
   
   Fuer alle Schluessel in vorlage, die nicht in eckigen Klammern stehen, wird ein gleicher
   Schluessel mit der gleichen Struktur in daten angelegt. In den Werten eines Schluessels wird
   ein Bereich an Zellen in einer Liste erwartet, die fuer diesen Schluessel eingelesen werden sollen
   (bspw. ["G14"] oder ["A7:A20"]). Falls die Zellen nicht als Text, sondern als Zahlenwerte
   eingelesen werden sollen, muss in einer Liste  Zellenbereich und Wertbereich angegeben werden,
   (bspw. ([["G14"], [0.0, 100.0]]).
   
   Wenn ein Schluessel in vorlage in eckigen Klammern steht, werden alle Eintraege in den Werten
   als Gruppe eingelesen. Das fuehrt dazu, dass bei Bereichen einzelne Werte von allen
   Gruppenmitgliedern uebersprungen werden, wenn fuer mindestens ein Gruppenmitglied kein gueltiger
   Wert existiert. Die Struktur mit dem Namen in eckigen Klammern wird nicht angelegt, sondern alle
   Gruppenmitglieder werden in der Elternstruktur gespeichert.
   """
   from .datenstruktur import Datenstruktur
   #
   for eintrag in vorlage.keys():
      if (isinstance(vorlage[eintrag], dict)):
         if (eintrag[0] == '['):
            _GruppeEinlesen(daten=daten, workbook=workbook, tabellentyp=tabellentyp,
               tabellenname=tabellenname, vorlage=vorlage[eintrag], gruppe=eintrag);
         else:
            unterstruktur = Datenstruktur();
            ParseXLSDaten(daten=unterstruktur, workbook=workbook, tabellentyp=tabellentyp,
               tabellenname=tabellenname, vorlage=vorlage[eintrag]);
            if (eintrag in daten):
               daten[eintrag].update(unterstruktur);
            else:
               daten.update([(eintrag, unterstruktur)]);
      #
      else:
         _EinzelreiheEinlesen(daten=daten, workbook=workbook, tabellentyp=tabellentyp,
            tabellenname=tabellenname, vorlage=vorlage, eintrag=eintrag);
#


# -------------------------------------------------------------------------------------------------
def _GruppeEinlesen(daten, workbook, tabellentyp, tabellenname, vorlage, gruppe):
   """Liest alle Eintraege in den Werten von vorlage mit dem Namen gruppe als Gruppe aus der
   Tabelle namens tabellenname des workbook vom Typ tabellentyp in die Datenstruktur daten ein. Alle
   Elemente aus vorlage ein gleich grosses Werteintervall haben. Wenn das erfuellt ist, werden fuer
   alle Elemente gleichmaessig Werte eingelesen (oder gemeinsam uebersprungen, wenn fuer mindestens
   ein Gruppenmitglied kein gueltiger Wert existiert). Die eingelesenen Daten sind gleich lang und
   werden direkt in daten gespeichert (gruppe wird nicht angelegt).
   """
   from .gleichungsloeser import LetzterIndexMitWertKleinerAls
   #
   eintraege = sorted(vorlage.keys());
   # Alle Gruppenmitglieder muessen die gleiche Intervallgroesse haben. Pruefe zusaetzlich den
   # zulaessigen Bereich aller Gruppenmitglieder
   zielbereiche = [];
   inhalte = [];
   grenzen = [];
   intervallgroesse = None;
   for idx_eintrag, eintrag in enumerate(eintraege):
      zellkoordinaten = ZellenbereichKoordinaten(workbook=workbook,
         tabellentyp=tabellentyp, tabellenname=tabellenname, bereich=vorlage[eintrag][0]);
      if (zellkoordinaten[0][0] is None):
         print('# Warnung: Ungueltige Zelle(n) ' + vorlage[eintrag][0] + ' in ' \
            + eintrag + ' - ignoriere ' + gruppe);
         return;
      #
      temp_intervallgroesse = len(zellkoordinaten);
      if (intervallgroesse is None):
         intervallgroesse = temp_intervallgroesse;
      else:
         if (temp_intervallgroesse != intervallgroesse):
            print('# Warnung: Intervalle in ' + gruppe + ' unterschiedlich gross - ignoriere Gruppe');
            return;
      #
      zielbereiche += [zellkoordinaten];
      if (len(vorlage[eintrag]) == 1):
         inhalte += ['text'];
         grenzen += [None];
      elif ((len(vorlage[eintrag]) == 2) or (len(vorlage[eintrag]) == 3)):
         inhalte += ['zahl'];
         grenzen += [vorlage[eintrag][1]];
      else:
         print('# Warnung: Ungueltige Laenge von ' + eintrag + ' - ignoriere ' + gruppe);
         return;
   #
   # Hier angekommen sind alle Intervalle gleich gross und im gueltigen Bereich. Ausserdem muss die
   # Tabellenseite existieren, da ansonsten schon abgebrochen worden waere
   tabellenseite = TabellenseiteAusgeben(workbook=workbook, tabellenname=tabellenname,
      tabellentyp=tabellentyp);
   wertliste = [[None for num in range(intervallgroesse)] for eintrag in eintraege];
   num_gute_werte = 0;
   verwerfe_Rest = False;
   idx_start = 0;
   idx_ende = intervallgroesse;
   for idx_intervall in range(intervallgroesse):
      for idx_eintrag, eintrag in enumerate(eintraege):
         idx_tempstart = 0;
         zellkoordinaten = zielbereiche[idx_eintrag];
         zeile, spalte = zellkoordinaten[idx_intervall];
         tempwert = _EinzelEintragEinlesen(tabellenseite=tabellenseite, zeile=zeile,
            spalte=spalte, name=eintrag, tabellentyp=tabellentyp, inhalt=inhalte[idx_eintrag],
            warnen=False);
         if (tempwert is None):
            if (num_gute_werte > 0):
               verwerfe_Rest = True;
               idx_ende = min(idx_ende, idx_intervall);
               break;
            else:
               idx_tempstart += 1;
               idx_start = max(idx_start, idx_tempstart);
               #
               continue;
         #
         wertliste[idx_eintrag][num_gute_werte] = tempwert;
      #
      if (verwerfe_Rest):
         break;
      #
      num_gute_werte += 1;
   #
   if (idx_start >= idx_ende-1):
      print('# Warnung: Konnte keine Werte fuer ' + gruppe + ' extrahieren');
      return;
   #
   # Listen zurechtschneiden, um bspw. noch vorhandene None-Eintraege zu entfernen
   for idx_liste in range(len(wertliste)):
      wertliste[idx_liste] = wertliste[idx_liste][idx_start:idx_ende];
   #
   # Wenn ein Wert aus der Gruppe eine Schnittgrenze hat, Startindex pruefen. Falls Startindex
   # groesser als null, alle Gruppenmitglieder erst ab diesem Startwert beginnen lassen
   idx_schnitt = 0;
   for idx_eintrag, eintrag in enumerate(eintraege):
      if (len(vorlage[eintrag]) == 3):
         # Nur fuer inhalt == 'zahl' mit vorhandenen grenzen zulaessig
         temp_idx = LetzterIndexMitWertKleinerAls(liste=wertliste[idx_eintrag],
            grenzwert=grenzen[idx_eintrag][0]);
         if (temp_idx is not None):
            idx_schnitt = max(idx_schnitt, temp_idx+1);
   #
   if (idx_schnitt > 0):
      if (idx_schnitt >= len(wertliste[0])-1):
         print('# Warnung: Konnte keine gueltigen Werte fuer ' + gruppe + ' extrahieren');
         return;
      #
      for idx_liste in range(len(wertliste)):
         wertliste[idx_liste] = wertliste[idx_liste][idx_schnitt:];
   #
   if ((idx_start + idx_schnitt > 0) or (idx_ende < intervallgroesse)):
      print('# Hinweis: Begrenze Werte in ' + gruppe + ' auf das Intervall [' \
         + str(idx_start + idx_schnitt) + ', ' + str(idx_ende) + ']');
   #
   for idx_eintrag, eintrag in enumerate(eintraege):
      GelesenenEintragHinzufuegen(datenbank=daten, name=eintrag,
         daten=wertliste[idx_eintrag], grenzen=grenzen[idx_eintrag]);
#


# -------------------------------------------------------------------------------------------------
def _EinzelreiheEinlesen(daten, workbook, tabellentyp, tabellenname, vorlage, eintrag):
   """Lese den eintrag aus vorlage aus der Tabelle namens tabellenname des workbook vom Typ
   tabellentyp in die Datenstruktur daten ein (sofern Bereich und ggfs. Werteintervall gueltig ist).
   """
   from .gleichungsloeser import LetzterIndexMitWertKleinerAls
   #
   if (len(vorlage[eintrag]) == 1):
      inhalt = 'text';
      grenzen = None;
   elif ((len(vorlage[eintrag]) == 2) or (len(vorlage[eintrag]) == 3)):
      inhalt = 'zahl';
      grenzen = vorlage[eintrag][1];
   else:
      print('# Warnung: Ungueltige Laenge von ' + eintrag);
      return;
   #
   zellkoordinaten = ZellenbereichKoordinaten(workbook=workbook,
      tabellentyp=tabellentyp, tabellenname=tabellenname, bereich=vorlage[eintrag][0]);
   if (zellkoordinaten[0][0] is None):
      print('# Warnung: Ungueltige Zelle(n) ' + vorlage[eintrag][0] + ' - Ignoriere ' + eintrag);
      return;
   #
   if (len(zellkoordinaten) == 1):
      EinzelEintragHinzufuegen(datenbank=daten, workbook=workbook,
         tabellenname=tabellenname, zeile=zellkoordinaten[0][0], spalte=zellkoordinaten[0][1],
         name=eintrag, tabellentyp=tabellentyp, inhalt=inhalt, grenzen=grenzen);
   else:
      werte = [];
      # Beendet den Lesevorgang bei einem None-Wert
      for zeile, spalte in zellkoordinaten:
         tempwert = EinzelEintragEinlesen(workbook=workbook, tabellenname=tabellenname,
            zeile=zeile, spalte=spalte, name=eintrag, tabellentyp=tabellentyp, inhalt=inhalt,
            warnen=False);
         if (tempwert is None):
            break;
         #
         werte += [tempwert];
      #
      if (len(werte) == 0):
         print('# Warnung: Leere Liste nach Einlesen ' + eintrag);
      else:
         idx_start = 0;
         if (len(vorlage[eintrag]) == 3):
            # Nur fuer inhalt == 'zahl' mit vorhandenen grenzen zulaessig
            temp_idx = LetzterIndexMitWertKleinerAls(liste=werte, grenzwert=grenzen[0]);
            if (temp_idx is not None):
               temp_idx += 1;
               idx_start = temp_idx;
         #
         if (idx_start > 0):
            print('# Hinweis: Schneide ungueltige Werte fuer ' + eintrag + ' bis nach Index ' \
               + str(idx_start) + ' ab');
         GelesenenEintragHinzufuegen(datenbank=daten, name=eintrag, daten=werte[idx_start:],
            grenzen=grenzen);
#
